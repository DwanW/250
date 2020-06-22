import psycopg2
from openpyxl import Workbook, load_workbook

conn = psycopg2.connect(database="postgres",
    user="jordzawada",
    password="123456789",
    host="farmersmarket-1.cqyj9z6amn2q.us-east-1.rds.amazonaws.com",
    port='5432')

cursor = conn.cursor()

def addFarmersRows():
    wb = load_workbook('./farmersmarkets.xlsx')
    maxID= cursor.execute("SELECT MAX(id) FROM farmers")
    maxID=[0][0]+1
    insert = "INSERT INTO farmers(id,name,lat,long) VALUES (%s,%s,%s,%s)"
    for row in wb['farmersmarkets'].iter_rows(min_row=2, max_row=wb['farmersmarkets'].max_row, values_only=True): 
        cursor.execute(insert, (maxID,row[3], row[15], row[16]))
        maxID=maxID+1


# addFarmersRows()

# addTest="INSERT INTO farmers(id,name,lat,long) VALUES (1,%s,%s,%s)"
# cursor.execute(addTest, ('name',51,120))

def printOut():
    cursor.execute("SELECT * FROM farmers")
    x=cursor.fetchall()
    print(x)

printOut()

conn.commit()
conn.close()











# import sqlite3
# from openpyxl import Workbook, load_workbook 


# connection = sqlite3.connect('data.db')

# cursor = connection.cursor()

# # MUST BE INTEGER
# # This is the only place where int vs INTEGER matters—in auto-incrementing columns
# create_table = "CREATE TABLE IF NOT EXISTS farmers (id INTEGER PRIMARY KEY, name text, lat int, long int)"
# cursor.execute(create_table)

# wb = load_workbook('./250/farmersmarkets.xlsx')
 
# # for row in wb['farmersmarkets']:
# #      # print(row[3].value)
# #     x = f"INSERT INTO farmers VALUES(null,{row[3].value},{row[15].value},{row[16].value}) "
# #     cursor.execute(x)
# def addRows():
#     insert = "INSERT into farmers VALUES(NULL,?,?,?)"
#     for row in wb['farmersmarkets'].iter_rows(min_row=2, max_row=wb['farmersmarkets'].max_row, values_only=True):   
#         cursor.execute(insert, (row[3], row[15], row[16]))
    
# # def viewTable():
# #     for row in cursor.execute("SELECT * FROM farmers"):
# #         print(list(row))

# # viewTable()
# connection.commit()

# connection.close()